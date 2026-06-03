#!/usr/bin/env python3
"""Export DOM tables from the active browser-use page to an XLSX workbook."""

from __future__ import annotations

import argparse
import datetime as dt
import json
import os
import re
import shutil
import subprocess
import sys
import time
from pathlib import Path
from typing import Any


START_MARKER = "AMAZON_TABLE_EXPORT_JSON_START"
END_MARKER = "AMAZON_TABLE_EXPORT_JSON_END"


SAFE_CLICK_JS = r"""
(() => {
  const safePattern = /(show\s*more|see\s*more|view\s*more|load\s*more|expand|展开|显示更多|查看更多|加载更多|更多)/i;
  const dangerPattern = /(submit|save|edit|delete|remove|buy|purchase|pay|confirm|send|提交|保存|编辑|删除|移除|购买|支付|确认|发送)/i;
  const visible = (el) => {
    const style = window.getComputedStyle(el);
    const rect = el.getBoundingClientRect();
    return style.display !== "none" && style.visibility !== "hidden" && rect.width > 0 && rect.height > 0;
  };

  const textOf = (el) => (el.innerText || el.getAttribute("aria-label") || el.getAttribute("title") || "").trim().replace(/\s+/g, " ");
  const clicked = window.__amazonTableExportClicked || [];
  const controls = Array.from(document.querySelectorAll("button, [role='button'], summary, a"));
  let count = 0;
  for (const el of controls) {
    if (count >= 30 || !visible(el)) continue;
    const text = textOf(el);
    if (!text || dangerPattern.test(text) || !safePattern.test(text)) continue;
    if (el.tagName === "A") {
      const href = el.getAttribute("href") || "";
      if (href && !href.startsWith("#") && !href.startsWith("javascript:")) continue;
    }
    try {
      el.scrollIntoView({ block: "center", inline: "nearest" });
      el.click();
      clicked.push(text.slice(0, 120));
      count += 1;
    } catch (_err) {
    }
  }
  window.__amazonTableExportClicked = clicked;
  return `AMAZON_TABLE_EXPORT_JSON_START${JSON.stringify({ clicked })}AMAZON_TABLE_EXPORT_JSON_END`;
})()
"""


EXTRACT_JS = r"""
(() => {
  const visible = (el) => {
    const style = window.getComputedStyle(el);
    const rect = el.getBoundingClientRect();
    return style.display !== "none" && style.visibility !== "hidden" && rect.width > 0 && rect.height > 0;
  };

  const textOf = (el) => (el.innerText || el.getAttribute("aria-label") || el.getAttribute("title") || "").trim().replace(/\s+/g, " ");

  const findContext = (el) => {
    const caption = el.querySelector("caption");
    if (caption && textOf(caption)) return textOf(caption);
    const aria = el.getAttribute("aria-label") || el.getAttribute("aria-labelledby");
    if (aria) return aria.trim();

    let node = el;
    for (let depth = 0; node && depth < 6; depth += 1, node = node.parentElement) {
      let prev = node.previousElementSibling;
      let guard = 0;
      while (prev && guard < 20) {
        const heading = prev.matches("h1,h2,h3,h4,h5,h6") ? prev : prev.querySelector("h1,h2,h3,h4,h5,h6");
        if (heading && textOf(heading)) return textOf(heading);
        prev = prev.previousElementSibling;
        guard += 1;
      }
    }

    const nearestHeading = document.querySelector("h1");
    return nearestHeading ? textOf(nearestHeading) : "";
  };

  const extractHtmlTable = (table, index) => {
    const occupied = [];
    const cells = [];
    let maxRow = 0;
    let maxCol = 0;
    const rows = Array.from(table.rows || []);

    rows.forEach((tr, r) => {
      occupied[r] = occupied[r] || [];
      let c = 0;
      Array.from(tr.cells || []).forEach((cell) => {
        while (occupied[r][c]) c += 1;
        const rowSpan = Math.max(1, parseInt(cell.getAttribute("rowspan") || "1", 10) || 1);
        const colSpan = Math.max(1, parseInt(cell.getAttribute("colspan") || "1", 10) || 1);
        for (let rr = 0; rr < rowSpan; rr += 1) {
          occupied[r + rr] = occupied[r + rr] || [];
          for (let cc = 0; cc < colSpan; cc += 1) occupied[r + rr][c + cc] = true;
        }
        cells.push({
          row: r + 1,
          col: c + 1,
          rowSpan,
          colSpan,
          text: textOf(cell),
          header: cell.tagName === "TH"
        });
        maxRow = Math.max(maxRow, r + rowSpan);
        maxCol = Math.max(maxCol, c + colSpan);
        c += colSpan;
      });
    });

    return {
      index,
      kind: "table",
      context: findContext(table),
      rows: maxRow,
      cols: maxCol,
      cells
    };
  };

  const extractRoleTable = (el, index) => {
    const rows = Array.from(el.querySelectorAll("[role='row']"));
    const cells = [];
    let maxCol = 0;
    rows.forEach((row, r) => {
      const rowCells = Array.from(row.querySelectorAll("[role='columnheader'], [role='rowheader'], [role='cell'], [role='gridcell']"));
      maxCol = Math.max(maxCol, rowCells.length);
      rowCells.forEach((cell, c) => {
        cells.push({
          row: r + 1,
          col: c + 1,
          rowSpan: 1,
          colSpan: 1,
          text: textOf(cell),
          header: /header/.test(cell.getAttribute("role") || "")
        });
      });
    });
    return {
      index,
      kind: el.getAttribute("role") || "role-table",
      context: findContext(el),
      rows: rows.length,
      cols: maxCol,
      cells
    };
  };

  const tables = [];
  const seen = new Set();
  Array.from(document.querySelectorAll("table")).forEach((table) => {
    if (!visible(table)) return;
    seen.add(table);
    const extracted = extractHtmlTable(table, tables.length + 1);
    if (extracted.cells.some((cell) => cell.text)) tables.push(extracted);
  });

  Array.from(document.querySelectorAll("[role='table'], [role='grid']")).forEach((el) => {
    if (seen.has(el) || !visible(el)) return;
    const extracted = extractRoleTable(el, tables.length + 1);
    if (extracted.cells.some((cell) => cell.text)) tables.push(extracted);
  });

  const payload = {
    title: document.title || "",
    url: location.href,
    clicked: window.__amazonTableExportClicked || [],
    tables
  };

  return `AMAZON_TABLE_EXPORT_JSON_START${JSON.stringify(payload)}AMAZON_TABLE_EXPORT_JSON_END`;
})()
"""


def ensure_openpyxl(install: bool) -> Any:
    try:
        import openpyxl  # type: ignore

        return openpyxl
    except ModuleNotFoundError:
        if not install:
            raise
        print("openpyxl is missing; installing it for the current user...", file=sys.stderr)
        subprocess.run([sys.executable, "-m", "pip", "install", "--user", "openpyxl"], check=True)
        import openpyxl  # type: ignore

        return openpyxl


def browser_use_base() -> list[str]:
    if shutil.which("browser-use"):
        return ["browser-use"]
    if shutil.which("uvx"):
        return ["uvx", "--python", "3.11", "browser-use"]
    raise RuntimeError("browser-use is not installed and uvx is not available")


def run_browser_use(args: list[str], session: str) -> str:
    completed = subprocess.run(
        browser_use_base() + ["--session", session] + args,
        check=True,
        text=True,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
    )
    if completed.stderr.strip():
        print(completed.stderr.strip(), file=sys.stderr)
    return completed.stdout


def parse_payload(output: str) -> dict[str, Any]:
    start = output.find(START_MARKER)
    end = output.find(END_MARKER)
    if start == -1 or end == -1 or end <= start:
        raise RuntimeError("Could not find table extraction JSON in browser-use eval output")
    raw = output[start + len(START_MARKER) : end]
    return json.loads(raw)


def prepare_page(session: str) -> None:
    run_browser_use(["eval", SAFE_CLICK_JS], session=session)
    for _ in range(20):
        try:
            run_browser_use(["scroll", "down", "--amount", "1000"], session=session)
        except subprocess.CalledProcessError:
            break
        time.sleep(0.08)
    run_browser_use(["eval", SAFE_CLICK_JS], session=session)
    for _ in range(8):
        try:
            run_browser_use(["scroll", "down", "--amount", "1200"], session=session)
        except subprocess.CalledProcessError:
            break
        time.sleep(0.08)
    run_browser_use(["eval", "(() => { window.scrollTo(0, 0); return 'ok'; })()"], session=session)


def downloads_dir() -> Path:
    return Path.home() / "Downloads"


def slug(value: str, fallback: str = "amazon-tables") -> str:
    cleaned = re.sub(r"[\\/:*?\"<>|\r\n\t]+", " ", value).strip()
    cleaned = re.sub(r"\s+", "-", cleaned)
    return (cleaned[:80] or fallback).strip("-") or fallback


def unique_sheet_name(base: str, used: set[str]) -> str:
    cleaned = re.sub(r"[\[\]:*?/\\]", " ", base).strip() or "Table"
    cleaned = re.sub(r"\s+", " ", cleaned)[:31]
    name = cleaned or "Table"
    counter = 2
    while name in used:
        suffix = f" {counter}"
        name = f"{cleaned[:31 - len(suffix)]}{suffix}"
        counter += 1
    used.add(name)
    return name


def open_file(path: Path) -> None:
    try:
        if sys.platform == "darwin":
            subprocess.Popen(["open", str(path)])
        elif os.name == "nt":
            os.startfile(str(path))  # type: ignore[attr-defined]
        else:
            subprocess.Popen(["xdg-open", str(path)])
    except Exception as exc:  # pragma: no cover - best effort convenience
        print(f"Could not open workbook automatically: {exc}", file=sys.stderr)


def autosize(ws: Any) -> None:
    from openpyxl.utils import get_column_letter

    for index, column_cells in enumerate(ws.columns, start=1):
        column = get_column_letter(index)
        max_len = 8
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, min(60, len(value) + 2))
        ws.column_dimensions[column].width = max_len


def write_workbook(payload: dict[str, Any], output: Path, install_deps: bool) -> Path:
    openpyxl = ensure_openpyxl(install_deps)
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    used_names: set[str] = set()

    thin = Side(style="thin", color="D0D7DE")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="EAECEF")
    index_fill = PatternFill("solid", fgColor="DDEBFF")
    header_font = Font(bold=True)
    wrap = Alignment(wrap_text=True, vertical="top")

    index_ws = wb.create_sheet(unique_sheet_name("Index", used_names))
    index_rows = [
        ["Page title", payload.get("title", "")],
        ["Source URL", payload.get("url", "")],
        ["Exported at", dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["Safe controls clicked", ", ".join(payload.get("clicked", []))],
        [],
        ["Sheet", "Context", "Rows", "Columns", "Kind"],
    ]
    for row in index_rows:
        index_ws.append(row)
    for cell in index_ws[1]:
        cell.fill = index_fill
        cell.font = header_font
    for cell in index_ws[6]:
        cell.fill = index_fill
        cell.font = header_font

    tables = payload.get("tables", [])
    for idx, table in enumerate(tables, start=1):
        context = table.get("context") or f"Table {idx}"
        sheet_name = unique_sheet_name(f"{idx} {context}", used_names)
        ws = wb.create_sheet(sheet_name)
        ws.freeze_panes = "A2"

        for cell in table.get("cells", []):
            row = int(cell.get("row", 1))
            col = int(cell.get("col", 1))
            row_span = max(1, int(cell.get("rowSpan", 1)))
            col_span = max(1, int(cell.get("colSpan", 1)))
            target = ws.cell(row=row, column=col, value=cell.get("text", ""))
            target.alignment = wrap
            target.border = border
            if cell.get("header") or row <= 2:
                target.fill = header_fill
                target.font = header_font
            if row_span > 1 or col_span > 1:
                ws.merge_cells(
                    start_row=row,
                    start_column=col,
                    end_row=row + row_span - 1,
                    end_column=col + col_span - 1,
                )
        for row in ws.iter_rows(min_row=1, max_row=max(1, int(table.get("rows", 1))), min_col=1, max_col=max(1, int(table.get("cols", 1)))):
            for cell in row:
                cell.alignment = wrap
                cell.border = border
        autosize(ws)

        index_ws.append([sheet_name, context, table.get("rows", 0), table.get("cols", 0), table.get("kind", "")])

    autosize(index_ws)
    index_ws.freeze_panes = "A7"
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    return output


def default_output(payload: dict[str, Any], output_dir: Path) -> Path:
    title = payload.get("title") or "amazon-tables"
    timestamp = dt.datetime.now().strftime("%Y%m%d-%H%M%S")
    return output_dir / f"{slug(title)}-{timestamp}.xlsx"


def main() -> int:
    parser = argparse.ArgumentParser(description="Export active browser-use page tables to Excel")
    parser.add_argument("--output", type=Path, help="Output XLSX path")
    parser.add_argument("--output-dir", type=Path, default=downloads_dir(), help="Directory for generated workbook")
    parser.add_argument("--open-file", action="store_true", help="Open the workbook after export")
    parser.add_argument("--no-install-deps", action="store_true", help="Do not auto-install missing Python packages")
    parser.add_argument("--session", default="default", help="browser-use session name")
    args = parser.parse_args()

    prepare_page(args.session)
    output = run_browser_use(["eval", EXTRACT_JS], session=args.session)
    payload = parse_payload(output)
    tables = payload.get("tables", [])
    if not tables:
        print("No DOM tables found on the active page.", file=sys.stderr)
        print(json.dumps({"title": payload.get("title"), "url": payload.get("url"), "tables": 0}, ensure_ascii=False))
        return 2

    workbook_path = args.output or default_output(payload, args.output_dir)
    write_workbook(payload, workbook_path, install_deps=not args.no_install_deps)
    if args.open_file:
        open_file(workbook_path)

    print(json.dumps({
        "workbook": str(workbook_path),
        "tables": len(tables),
        "title": payload.get("title", ""),
        "url": payload.get("url", ""),
    }, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
